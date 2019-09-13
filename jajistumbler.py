# -*- coding: utf-8 -*-

import csv
import os
from openpyxl import *

class place(object):
	def __init__(self):
		self._nameplace = ""
		self._listwifi = []
	@property
	def nameplace(self):
		return self._nameplace

	@nameplace.setter
	def nameplace(self, n):
		self._nameplace = n
		return self._nameplace
	@property
	def listwifi(self):
		return self._listwifi

	@listwifi.setter
	def listwifi(self, l):
		self._listwifi = self._listwifi+[l]
		return self._listwifi

class signalwifi(object):
	def __init__(self):
		self._name = ""
		self._bssid = []
		self._signalintensity = ""
		self._filename = ""
	@property
	def name(self):
		return self._name

	@name.setter
	def name(self, n):
		self._name = n
		return self._name

	@property
	def bssid(self):
		return self._bssid

	@bssid.setter
	def bssid(self, b):
		self._bssid = self._bssid+[b]
		return self._bssid

	@property
	def signalintensity(self):
		return self._signalintensity

	@signalintensity.setter
	def signalintensity(self, s):
		self._signalintensity = s
		return self._signalintensity

	@property
	def filename(self):
		return self._filename

	@filename.setter
	def filename(self, f):
		self._filename = f
		return self._filename

def triallwifi(listwifiall):
	listtmp = []
	for current in listwifiall:
		flag = False
		if listtmp == []:
			listtmp.append(current)
		for currenttmp in listtmp:
			if currenttmp.bssid[0] == current.bssid[0]:
				flag = True
				if int(currenttmp.signalintensity) < int(current.signalintensity):
					currenttmp.signalintensity = current.signalintensity
		if flag == False:
			listtmp.append(current)
	return listtmp

def checkwifiname(name, listwifi):
	for count, current in enumerate(listwifi):
		if current.name == name:
			return count
	return -1

listfile = []
listplace = []
listallwifi = []
listwififort = []
listwifibssid = []

for currentfile in os.listdir("./csv"):
	if currentfile.endswith(".csv"):
		listfile.append(os.path.join("./csv", currentfile))

for currentfile in listfile:
	newPlace = place()
	newPlace.nameplace = currentfile.split('.csv')[0].split('/')[2]
	listbssid = []
	with open(currentfile) as csvfile:
		readCSV = csv.reader(csvfile, delimiter=',')
		for row in readCSV:
			if(row[1] == "BSSID"):
				continue
			if (row[1] in listbssid):
				continue
			listbssid.append(row[1])
			newWIFI = signalwifi()
			newWIFI.name = row[0]
			if row[1] not in newWIFI.bssid:
				newWIFI.bssid = row[1]
			newWIFI.signalintensity = row[6]
			newWIFI.filename = currentfile.split('.csv')[0].split('/')[2]
			newPlace.listwifi = newWIFI
			idx = checkwifiname(row[0], listwifibssid)
			if idx == -1:
				listwifibssid.append(newWIFI)
			else:
				if row[1] not in listwifibssid[idx].bssid:
					listwifibssid[idx].bssid = row[1]
			if int(row[6]) >= -50:
				idx = checkwifiname(row[0], listwififort)
				if idx == -1:
					listwififort.append(newWIFI)
				else:
					if row[1] not in listwififort[idx].bssid:
						listwififort[idx].bssid = row[1]
			listallwifi.append(newWIFI)
	listplace.append(newPlace)



wb = Workbook()
wsnew = wb.create_sheet("WIFI Unique")

currentline = 2
wsnew['A1'] = "Nom WI-FI"
wsnew['B1'] = "BSSID"
wsnew['C1'] = "Intensite MAX"
wsnew['D1'] = "Fichier"
for currentwifi in triallwifi(listallwifi):
	wsnew['A'+str(currentline)] = currentwifi.name
	wsnew['B'+str(currentline)] = currentwifi.bssid[0]
	wsnew['C'+str(currentline)] = currentwifi.signalintensity
	wsnew['D'+str(currentline)] = currentwifi.filename
	currentline+=1

wsnew = wb.create_sheet("WIFI Signal Fort (>=-50db)")
wsnew['A1'] = "Nom WI-FI"
wsnew['B1'] = "BSSID"
wsnew['C1'] = "Nombre total de BSSID"
currentline = 2
for currentwifi in listwififort:
	linebssid = ""
	for current in currentwifi.bssid:
		if linebssid == "":
			linebssid = current
		else:
			linebssid = linebssid+', '+current
	wsnew['A'+str(currentline)] = currentwifi.name
	wsnew['B'+str(currentline)] = linebssid
	wsnew['C'+str(currentline)] = len(currentwifi.bssid)
	currentline+=1

wsnew = wb.create_sheet("WIFI BSSID")
wsnew['A1'] = "Nom WI-FI"
wsnew['B1'] = "BSSID"
wsnew['C1'] = "Nombre total de BSSID"
currentline = 2
for currentwifi in listwifibssid:
	linebssid = ""
	for current in currentwifi.bssid:
		if linebssid == "":
			linebssid = current
		else:
			linebssid = linebssid+', '+current
	wsnew['A'+str(currentline)] = currentwifi.name
	wsnew['B'+str(currentline)] = linebssid
	wsnew['C'+str(currentline)] = len(currentwifi.bssid)
	currentline+=1

for current in listplace:
	wsnew = wb.create_sheet(current.nameplace)
	wsnew['A1'] = "Nom WI-FI"
	wsnew['B1'] = "BSSID"
	wsnew['C1'] = "Intensite MAX"
	currentline = 2
	for currentwifi in current.listwifi:
		wsnew['A'+str(currentline)] = currentwifi.name
		wsnew['B'+str(currentline)] = currentwifi.bssid[0]
		wsnew['C'+str(currentline)] = currentwifi.signalintensity
		currentline+=1
wb.remove(wb['Sheet'])
wb.save('Jajesult.xlsx')